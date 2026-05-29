"""
Parser voor Google Ads Excel/CSV exports.

Ondersteunde exportformaten:
- Campagne-rapport (Campaign / Campagne)
- Advertentiegroep-rapport (Ad group / Advertentiegroep)
- Zoekwoord-rapport (Keyword / Zoekwoord) — inclusief kwaliteitsscore
- Zoekterm-rapport (Search term / Zoekterm)
- Zowel .xlsx als .csv, EN en NL kolomnamen
- CTR als % of decimaal, kosten met of zonder €-teken

Exportinstructies Google Ads:
1. Open Google Ads > Rapporten (of gebruik de tabelweergave per niveau)
2. Selecteer het gewenste niveau: Campagnes / Advertentiegroepen / Zoekwoorden / Zoektermen
3. Voeg kolommen toe: Klikken, Vertoningen, CTR, Gem. CPC, Kosten, Conversies,
   Conversieratio, Kosten/conversie (en Kwaliteitsscore voor zoekwoorden)
4. Exporteer als Excel (.xlsx) of CSV
5. Herhaal voor de vergelijkingsperiode en geef mee via --vergelijk
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Datamodellen
# ---------------------------------------------------------------------------

@dataclass
class CampaignRow:
    campaign: str
    clicks: int
    impressions: int
    ctr: float          # percentage (bijv. 4.55)
    avg_cpc: float      # euro
    cost: float         # euro
    conversions: float
    conv_rate: float    # percentage
    cost_per_conv: float
    budget: float = 0.0
    conv_value: float = 0.0      # conversiewaarde (€)
    impression_share: float = 0.0       # vertoningsaandeel (%)
    lost_is_budget: float = 0.0         # IS-verlies door budget (%)
    lost_is_rank: float = 0.0           # IS-verlies door ad rank (%)


@dataclass
class AdGroupRow:
    campaign: str
    ad_group: str
    clicks: int
    impressions: int
    ctr: float
    avg_cpc: float
    cost: float
    conversions: float
    conv_rate: float
    cost_per_conv: float
    conv_value: float = 0.0
    impression_share: float = 0.0
    lost_is_budget: float = 0.0
    lost_is_rank: float = 0.0


@dataclass
class KeywordRow:
    campaign: str
    ad_group: str
    keyword: str
    match_type: str
    clicks: int
    impressions: int
    ctr: float
    avg_cpc: float
    cost: float
    conversions: float
    conv_rate: float
    cost_per_conv: float
    quality_score: int = 0
    conv_value: float = 0.0
    impression_share: float = 0.0
    lost_is_budget: float = 0.0
    lost_is_rank: float = 0.0


@dataclass
class SearchTermRow:
    search_term: str
    keyword: str
    campaign: str
    ad_group: str
    match_type: str
    clicks: int
    impressions: int
    ctr: float
    avg_cpc: float
    cost: float
    conversions: float
    conv_rate: float
    cost_per_conv: float
    conv_value: float = 0.0


@dataclass
class AuctionInsightRow:
    """Rij uit het veilinginzichten-rapport."""
    competitor: str          # domeinnaam concurrent
    impression_share: float  # vertoningsaandeel concurrent (%)
    overlap_rate: float      # overlap rate (%)
    position_above_rate: float  # hoe vaak concurrent boven jou staat (%)
    top_of_page_rate: float  # % vertoningen bovenaan pagina
    abs_top_of_page_rate: float  # % vertoningen op absolute toppositie
    outranking_share: float  # hoe vaak jij boven concurrent staat (%)
    campaign: str = ""
    ad_group: str = ""


@dataclass
class AdsData:
    """Container voor alle geparsede Google Ads data uit één of meerdere bestanden."""
    campaigns: list[CampaignRow]
    ad_groups: list[AdGroupRow]
    keywords: list[KeywordRow]
    search_terms: list[SearchTermRow]
    auction_insights: list[AuctionInsightRow]
    source_files: list[str]


# ---------------------------------------------------------------------------
# Kolom-normalisatie
# ---------------------------------------------------------------------------

_COL_MAP = {
    # Campagne
    "campaign": "campaign",
    "campagne": "campaign",
    "campaign name": "campaign",
    "campagnenaam": "campaign",
    # Advertentiegroep
    "ad group": "ad_group",
    "advertentiegroep": "ad_group",
    "ad group name": "ad_group",
    "advertentiegroepnaam": "ad_group",
    # Zoekwoord
    "keyword": "keyword",
    "zoekwoord": "keyword",
    "keyword text": "keyword",
    "search keyword": "keyword",
    # Zoekterm
    "search term": "search_term",
    "zoekterm": "search_term",
    "search query": "search_term",
    "zoekopdracht": "search_term",
    # Match type
    "match type": "match_type",
    "zoektype": "match_type",
    "keyword match type": "match_type",
    "zoekwoordtype": "match_type",
    # Metrics
    "clicks": "clicks",
    "klikken": "clicks",
    "impressions": "impressions",
    "vertoningen": "impressions",
    "ctr": "ctr",
    "avg. cpc": "avg_cpc",
    "gem. cpc": "avg_cpc",
    "gemiddelde cpc": "avg_cpc",
    "average cpc": "avg_cpc",
    "cost": "cost",
    "kosten": "cost",
    "spend": "cost",
    "uitgaven": "cost",
    "conversions": "conversions",
    "conversies": "conversions",
    "conv.": "conversions",
    "conv. rate": "conv_rate",
    "conversieratio": "conv_rate",
    "conversion rate": "conv_rate",
    "cost / conv.": "cost_per_conv",
    "kosten/conversie": "cost_per_conv",
    "cost per conversion": "cost_per_conv",
    "cpa": "cost_per_conv",
    "quality score": "quality_score",
    "kwaliteitsscore": "quality_score",
    "qual. score": "quality_score",
    # Budget
    "budget": "budget",
    "daily budget": "budget",
    "dagelijks budget": "budget",
    # Conversiewaarde / ROAS
    "conversion value": "conv_value",
    "conversiewaarde": "conv_value",
    "conv. value": "conv_value",
    "all conv. value": "conv_value",
    "alle conv.waarde": "conv_value",
    "value": "conv_value",
    # Vertoningsaandeel
    "search impr. share": "impression_share",
    "zoekvertoningsaandeel": "impression_share",
    "impr. share": "impression_share",
    "vertoningsaandeel": "impression_share",
    "impression share": "impression_share",
    "search impression share": "impression_share",
    # IS-verlies budget
    "search lost is (budget)": "lost_is_budget",
    "search lost impr. share (budget)": "lost_is_budget",
    "zoek verloren vs. (budget)": "lost_is_budget",
    "lost is (budget)": "lost_is_budget",
    # IS-verlies rank
    "search lost is (rank)": "lost_is_rank",
    "search lost impr. share (rank)": "lost_is_rank",
    "zoek verloren vs. (rang)": "lost_is_rank",
    "lost is (rank)": "lost_is_rank",
    # Auction Insights kolommen
    "display url domain": "competitor",
    "domein": "competitor",
    "competitor": "competitor",
    "overlap rate": "overlap_rate",
    "overlappingspercentage": "overlap_rate",
    "position above rate": "position_above_rate",
    "percentage hogere positie": "position_above_rate",
    "top of page rate": "top_of_page_rate",
    "percentage bovenaan pagina": "top_of_page_rate",
    "abs. top of page rate": "abs_top_of_page_rate",
    "percentage absolute toppositie": "abs_top_of_page_rate",
    "outranking share": "outranking_share",
    "percentage hoger gerangschikt": "outranking_share",
}


def _normalize_header(raw: str) -> str:
    return _COL_MAP.get(raw.strip().lower(), raw.strip().lower())


def _parse_float(value) -> float:
    """Zet waarde om naar float. Verwerkt €, %, komma als decimaalscheiding."""
    v = str(value).strip().replace("€", "").replace(",", ".").replace("%", "").replace("\xa0", "").strip()
    # Verwijder duizendtalscheiding (punt als scheider bij grote getallen)
    # Detecteer: als er meerdere punten zijn, zijn de eerste duizendtalscheiders
    parts = v.split(".")
    if len(parts) > 2:
        v = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return float(v) if v else 0.0
    except ValueError:
        return 0.0


def _parse_ctr(value) -> float:
    """CTR: als < 1, vermenigvuldig met 100 (soms 0.0455 i.p.v. 4.55%)."""
    f = _parse_float(value)
    return round(f * 100, 4) if 0 < f < 1 else round(f, 4)


def _parse_pct(value) -> float:
    """Conversieratio: zelfde logica als CTR."""
    return _parse_ctr(value)


def _parse_int(value) -> int:
    return int(_parse_float(value))


# ---------------------------------------------------------------------------
# Sheet/CSV lezen naar lijst van dicts
# ---------------------------------------------------------------------------

def _rows_from_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        return [{_normalize_header(k): v for k, v in row.items()} for row in reader]


def _rows_from_sheet(sheet) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(str(h) if h is not None else "") for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        result.append({headers[i]: (str(row[i]) if row[i] is not None else "") for i in range(len(headers))})
    return result


# ---------------------------------------------------------------------------
# Type-detectie per set rijen
# ---------------------------------------------------------------------------

def _detect_sheet_type(rows: list[dict]) -> str:
    """Geeft 'campaigns', 'ad_groups', 'keywords', 'search_terms', 'auction_insights', of 'unknown'."""
    if not rows:
        return "unknown"
    keys = set(rows[0].keys())
    has_search_term = "search_term" in keys
    has_keyword = "keyword" in keys
    has_ad_group = "ad_group" in keys
    has_campaign = "campaign" in keys
    has_competitor = "competitor" in keys

    if has_competitor:
        return "auction_insights"
    if has_search_term:
        return "search_terms"
    if has_keyword:
        return "keywords"
    if has_ad_group:
        return "ad_groups"
    if has_campaign:
        return "campaigns"
    return "unknown"


# ---------------------------------------------------------------------------
# Rijen omzetten naar dataclasses
# ---------------------------------------------------------------------------

def _parse_is(value) -> float:
    """Vertoningsaandeel: kan '--' zijn als data ontbreekt, anders % of decimaal."""
    v = str(value).strip()
    if v in ("--", "-", ""):
        return 0.0
    return _parse_ctr(v)


def _to_campaign_rows(rows: list[dict]) -> list[CampaignRow]:
    result = []
    for r in rows:
        try:
            result.append(CampaignRow(
                campaign=r.get("campaign", "").strip(),
                clicks=_parse_int(r.get("clicks", "0")),
                impressions=_parse_int(r.get("impressions", "0")),
                ctr=_parse_ctr(r.get("ctr", "0")),
                avg_cpc=_parse_float(r.get("avg_cpc", "0")),
                cost=_parse_float(r.get("cost", "0")),
                conversions=_parse_float(r.get("conversions", "0")),
                conv_rate=_parse_pct(r.get("conv_rate", "0")),
                cost_per_conv=_parse_float(r.get("cost_per_conv", "0")),
                budget=_parse_float(r.get("budget", "0")),
                conv_value=_parse_float(r.get("conv_value", "0")),
                impression_share=_parse_is(r.get("impression_share", "0")),
                lost_is_budget=_parse_is(r.get("lost_is_budget", "0")),
                lost_is_rank=_parse_is(r.get("lost_is_rank", "0")),
            ))
        except (ValueError, TypeError):
            continue
    return [c for c in result if c.campaign]


def _to_ad_group_rows(rows: list[dict]) -> list[AdGroupRow]:
    result = []
    for r in rows:
        try:
            result.append(AdGroupRow(
                campaign=r.get("campaign", "").strip(),
                ad_group=r.get("ad_group", "").strip(),
                clicks=_parse_int(r.get("clicks", "0")),
                impressions=_parse_int(r.get("impressions", "0")),
                ctr=_parse_ctr(r.get("ctr", "0")),
                avg_cpc=_parse_float(r.get("avg_cpc", "0")),
                cost=_parse_float(r.get("cost", "0")),
                conversions=_parse_float(r.get("conversions", "0")),
                conv_rate=_parse_pct(r.get("conv_rate", "0")),
                cost_per_conv=_parse_float(r.get("cost_per_conv", "0")),
                conv_value=_parse_float(r.get("conv_value", "0")),
                impression_share=_parse_is(r.get("impression_share", "0")),
                lost_is_budget=_parse_is(r.get("lost_is_budget", "0")),
                lost_is_rank=_parse_is(r.get("lost_is_rank", "0")),
            ))
        except (ValueError, TypeError):
            continue
    return [a for a in result if a.ad_group]


def _to_keyword_rows(rows: list[dict]) -> list[KeywordRow]:
    result = []
    for r in rows:
        try:
            qs_raw = r.get("quality_score", "0")
            qs = _parse_int(qs_raw) if qs_raw and str(qs_raw).strip() not in ("", "--", "-") else 0
            result.append(KeywordRow(
                campaign=r.get("campaign", "").strip(),
                ad_group=r.get("ad_group", "").strip(),
                keyword=r.get("keyword", "").strip(),
                match_type=r.get("match_type", "").strip(),
                clicks=_parse_int(r.get("clicks", "0")),
                impressions=_parse_int(r.get("impressions", "0")),
                ctr=_parse_ctr(r.get("ctr", "0")),
                avg_cpc=_parse_float(r.get("avg_cpc", "0")),
                cost=_parse_float(r.get("cost", "0")),
                conversions=_parse_float(r.get("conversions", "0")),
                conv_rate=_parse_pct(r.get("conv_rate", "0")),
                cost_per_conv=_parse_float(r.get("cost_per_conv", "0")),
                quality_score=qs,
                conv_value=_parse_float(r.get("conv_value", "0")),
                impression_share=_parse_is(r.get("impression_share", "0")),
                lost_is_budget=_parse_is(r.get("lost_is_budget", "0")),
                lost_is_rank=_parse_is(r.get("lost_is_rank", "0")),
            ))
        except (ValueError, TypeError):
            continue
    return [k for k in result if k.keyword]


def _to_search_term_rows(rows: list[dict]) -> list[SearchTermRow]:
    result = []
    for r in rows:
        try:
            result.append(SearchTermRow(
                search_term=r.get("search_term", "").strip(),
                keyword=r.get("keyword", "").strip(),
                campaign=r.get("campaign", "").strip(),
                ad_group=r.get("ad_group", "").strip(),
                match_type=r.get("match_type", "").strip(),
                clicks=_parse_int(r.get("clicks", "0")),
                impressions=_parse_int(r.get("impressions", "0")),
                ctr=_parse_ctr(r.get("ctr", "0")),
                avg_cpc=_parse_float(r.get("avg_cpc", "0")),
                cost=_parse_float(r.get("cost", "0")),
                conversions=_parse_float(r.get("conversions", "0")),
                conv_rate=_parse_pct(r.get("conv_rate", "0")),
                cost_per_conv=_parse_float(r.get("cost_per_conv", "0")),
                conv_value=_parse_float(r.get("conv_value", "0")),
            ))
        except (ValueError, TypeError):
            continue
    return [s for s in result if s.search_term]


def _to_auction_insight_rows(rows: list[dict]) -> list[AuctionInsightRow]:
    result = []
    for r in rows:
        try:
            result.append(AuctionInsightRow(
                competitor=r.get("competitor", "").strip(),
                impression_share=_parse_is(r.get("impression_share", "0")),
                overlap_rate=_parse_is(r.get("overlap_rate", "0")),
                position_above_rate=_parse_is(r.get("position_above_rate", "0")),
                top_of_page_rate=_parse_is(r.get("top_of_page_rate", "0")),
                abs_top_of_page_rate=_parse_is(r.get("abs_top_of_page_rate", "0")),
                outranking_share=_parse_is(r.get("outranking_share", "0")),
                campaign=r.get("campaign", "").strip(),
                ad_group=r.get("ad_group", "").strip(),
            ))
        except (ValueError, TypeError):
            continue
    return [a for a in result if a.competitor]


# ---------------------------------------------------------------------------
# Publieke API
# ---------------------------------------------------------------------------

def parse_file(path: str | Path) -> AdsData:
    """
    Parseer één Google Ads Excel of CSV exportbestand.

    Ondersteunt:
    - .csv  — één sheet (campagne, advertentiegroep, zoekwoord of zoekterm)
    - .xlsx — één of meerdere sheets (auto-detect per sheet)
    """
    p = Path(path)
    campaigns: list[CampaignRow] = []
    ad_groups: list[AdGroupRow] = []
    keywords: list[KeywordRow] = []
    search_terms: list[SearchTermRow] = []
    auction_insights: list[AuctionInsightRow] = []

    if p.suffix.lower() == ".csv":
        rows = _rows_from_csv(p)
        sheet_type = _detect_sheet_type(rows)
        if sheet_type == "campaigns":
            campaigns = _to_campaign_rows(rows)
        elif sheet_type == "ad_groups":
            ad_groups = _to_ad_group_rows(rows)
        elif sheet_type == "keywords":
            keywords = _to_keyword_rows(rows)
        elif sheet_type == "search_terms":
            search_terms = _to_search_term_rows(rows)
        elif sheet_type == "auction_insights":
            auction_insights = _to_auction_insight_rows(rows)

    elif p.suffix.lower() in (".xlsx", ".xls"):
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl is vereist voor .xlsx bestanden: pip install openpyxl")
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = _rows_from_sheet(ws)
            sheet_type = _detect_sheet_type(rows)
            if sheet_type == "campaigns":
                campaigns.extend(_to_campaign_rows(rows))
            elif sheet_type == "ad_groups":
                ad_groups.extend(_to_ad_group_rows(rows))
            elif sheet_type == "keywords":
                keywords.extend(_to_keyword_rows(rows))
            elif sheet_type == "search_terms":
                search_terms.extend(_to_search_term_rows(rows))
            elif sheet_type == "auction_insights":
                auction_insights.extend(_to_auction_insight_rows(rows))
        wb.close()
    else:
        raise ValueError(f"Niet-ondersteund bestandstype: {p.suffix}. Gebruik .csv of .xlsx")

    return AdsData(
        campaigns=campaigns,
        ad_groups=ad_groups,
        keywords=keywords,
        search_terms=search_terms,
        auction_insights=auction_insights,
        source_files=[str(p)],
    )


def parse_files(paths: list[str | Path]) -> AdsData:
    """
    Parseer meerdere Google Ads exportbestanden en combineer ze tot één AdsData object.
    Handig als campagnes, zoekwoorden en zoektermen als aparte exports zijn opgeslagen.
    """
    combined = AdsData(campaigns=[], ad_groups=[], keywords=[], search_terms=[], auction_insights=[], source_files=[])
    for path in paths:
        data = parse_file(path)
        combined.campaigns.extend(data.campaigns)
        combined.ad_groups.extend(data.ad_groups)
        combined.keywords.extend(data.keywords)
        combined.search_terms.extend(data.search_terms)
        combined.auction_insights.extend(data.auction_insights)
        combined.source_files.extend(data.source_files)
    return combined


def validate(data: AdsData) -> list[str]:
    """Geeft een lijst van waarschuwingen als data onvolledig of leeg is."""
    warnings = []
    if not data.campaigns and not data.keywords and not data.ad_groups:
        warnings.append("Geen campagne-, advertentiegroep- of zoekwoorddata gevonden. Controleer de kolomnamen.")
    if not data.campaigns:
        warnings.append("Geen campagnedata gevonden — campagneniveau-analyses niet mogelijk.")
    if not data.keywords:
        warnings.append("Geen zoekwoorddata gevonden — zoekwoordanalyses niet mogelijk.")
    if not data.search_terms:
        warnings.append("Geen zoektermdata gevonden — zoektermanalyses (uitsluitingszoekwoorden, kansen) niet mogelijk.")
    return warnings
