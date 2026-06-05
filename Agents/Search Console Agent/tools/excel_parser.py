"""
Parser voor Google Search Console Excel/CSV exports.

Ondersteunde exportformaten:
- GSC Performance > Queries (kolom: "Top queries" of "Query")
- GSC Performance > Pages (kolom: "Top pages" of "Page" of "Pagina")
- Gecombineerde sheet met Query + Page (voor cannibalisatiedetectie)
- Zowel .xlsx als .csv, EN en NL kolomnamen, CTR als % of decimaal
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Datamodellen
# ---------------------------------------------------------------------------

@dataclass
class QueryRow:
    query: str
    clicks: int
    impressions: int
    ctr: float        # percentage (bijv. 4.55)
    position: float


@dataclass
class PageRow:
    page: str
    clicks: int
    impressions: int
    ctr: float
    position: float


@dataclass
class QueryPageRow:
    """Gecombineerde rij: zoekterm + landingspagina — voor cannibalisatiedetectie."""
    query: str
    page: str
    clicks: int
    impressions: int
    ctr: float
    position: float


@dataclass
class GSCData:
    """Container voor alle geparsede GSC data uit één of meerdere bestanden."""
    queries: list[QueryRow]
    pages: list[PageRow]
    query_pages: list[QueryPageRow]   # alleen aanwezig als gecombineerde sheet beschikbaar is
    source_files: list[str]


# ---------------------------------------------------------------------------
# Kolom-normalisatie
# ---------------------------------------------------------------------------

# Mapping van mogelijke kolomnamen → interne sleutelnaam
_COL_MAP = {
    # Query kolom
    "top queries": "query",
    "query": "query",
    "zoekopdracht": "query",
    "zoekopdrachten": "query",
    "top zoekopdrachten": "query",
    # Page kolom
    "top pages": "page",
    "page": "page",
    "pagina": "page",
    "pagina's": "page",
    "top pagina's": "page",
    "top paginas": "page",
    # Metrics
    "clicks": "clicks",
    "klikken": "clicks",
    "impressions": "impressions",
    "vertoningen": "impressions",
    "ctr": "ctr",
    "position": "position",
    "positie": "position",
    "avg. position": "position",
    "gemiddelde positie": "position",
}


def _normalize_header(raw: str) -> str:
    return _COL_MAP.get(raw.strip().lower(), raw.strip().lower())


def _parse_float(value: str) -> float:
    """Zet string om naar float. Verwerkt ook % en komma als decimaalscheiding."""
    v = str(value).strip().replace(",", ".").replace("%", "").strip()
    return float(v) if v else 0.0


def _parse_ctr(value: str) -> float:
    """CTR: als het een waarde < 1 is, vermenigvuldig met 100 (GSC geeft soms 0.0455)."""
    raw = str(value).strip().replace(",", ".").replace("%", "").strip()
    f = float(raw) if raw else 0.0
    return round(f * 100, 4) if f < 1.0 else round(f, 4)


# ---------------------------------------------------------------------------
# Sheet/CSV lezen naar lijst van dicts
# ---------------------------------------------------------------------------

def _rows_from_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        normalized = []
        for row in reader:
            normalized.append({_normalize_header(k): v for k, v in row.items()})
        return normalized


def _rows_from_sheet(sheet) -> list[dict]:
    """Lees een openpyxl worksheet naar lijst van dicts."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(str(h) if h is not None else "") for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        result.append({headers[i]: (str(row[i]) if row[i] is not None else "") for i in range(len(headers))})
    return result


# ---------------------------------------------------------------------------
# Type-detectie per set rijen
# ---------------------------------------------------------------------------

def _detect_sheet_type(rows: list[dict]) -> str:
    """Geeft 'queries', 'pages', 'query_pages', of 'unknown'."""
    if not rows:
        return "unknown"
    keys = set(rows[0].keys())
    has_query = "query" in keys
    has_page = "page" in keys
    if has_query and has_page:
        return "query_pages"
    if has_query:
        return "queries"
    if has_page:
        return "pages"
    return "unknown"


# ---------------------------------------------------------------------------
# Rijen omzetten naar dataclasses
# ---------------------------------------------------------------------------

def _to_query_rows(rows: list[dict]) -> list[QueryRow]:
    result = []
    for r in rows:
        try:
            result.append(QueryRow(
                query=r.get("query", "").strip(),
                clicks=int(_parse_float(r.get("clicks", "0"))),
                impressions=int(_parse_float(r.get("impressions", "0"))),
                ctr=_parse_ctr(r.get("ctr", "0")),
                position=round(_parse_float(r.get("position", "0")), 1),
            ))
        except (ValueError, TypeError):
            continue
    return [q for q in result if q.query]


def _to_page_rows(rows: list[dict]) -> list[PageRow]:
    result = []
    for r in rows:
        try:
            result.append(PageRow(
                page=r.get("page", "").strip(),
                clicks=int(_parse_float(r.get("clicks", "0"))),
                impressions=int(_parse_float(r.get("impressions", "0"))),
                ctr=_parse_ctr(r.get("ctr", "0")),
                position=round(_parse_float(r.get("position", "0")), 1),
            ))
        except (ValueError, TypeError):
            continue
    return [p for p in result if p.page]


def _to_query_page_rows(rows: list[dict]) -> list[QueryPageRow]:
    result = []
    for r in rows:
        try:
            result.append(QueryPageRow(
                query=r.get("query", "").strip(),
                page=r.get("page", "").strip(),
                clicks=int(_parse_float(r.get("clicks", "0"))),
                impressions=int(_parse_float(r.get("impressions", "0"))),
                ctr=_parse_ctr(r.get("ctr", "0")),
                position=round(_parse_float(r.get("position", "0")), 1),
            ))
        except (ValueError, TypeError):
            continue
    return [qp for qp in result if qp.query and qp.page]


# ---------------------------------------------------------------------------
# Publieke API
# ---------------------------------------------------------------------------

def parse_file(path: str | Path) -> GSCData:
    """
    Parseer één Excel of CSV bestand met GSC exportdata.

    Ondersteunt:
    - .csv  — één sheet (queries óf pages)
    - .xlsx — één of meerdere sheets (auto-detect per sheet)
    """
    p = Path(path)
    queries: list[QueryRow] = []
    pages: list[PageRow] = []
    query_pages: list[QueryPageRow] = []

    if p.suffix.lower() == ".csv":
        rows = _rows_from_csv(p)
        sheet_type = _detect_sheet_type(rows)
        if sheet_type == "queries":
            queries = _to_query_rows(rows)
        elif sheet_type == "pages":
            pages = _to_page_rows(rows)
        elif sheet_type == "query_pages":
            query_pages = _to_query_page_rows(rows)

    elif p.suffix.lower() in (".xlsx", ".xls"):
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl is vereist voor .xlsx bestanden: pip install openpyxl")
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = _rows_from_sheet(ws)
            sheet_type = _detect_sheet_type(rows)
            if sheet_type == "queries":
                queries.extend(_to_query_rows(rows))
            elif sheet_type == "pages":
                pages.extend(_to_page_rows(rows))
            elif sheet_type == "query_pages":
                query_pages.extend(_to_query_page_rows(rows))
        wb.close()
    else:
        raise ValueError(f"Niet-ondersteund bestandstype: {p.suffix}. Gebruik .csv of .xlsx")

    return GSCData(
        queries=queries,
        pages=pages,
        query_pages=query_pages,
        source_files=[str(p)],
    )


def parse_files(paths: list[str | Path]) -> GSCData:
    """
    Parseer meerdere GSC exportbestanden en combineer ze tot één GSCData object.
    Handig als queries en pages als aparte exports zijn opgeslagen.
    """
    combined = GSCData(queries=[], pages=[], query_pages=[], source_files=[])
    for path in paths:
        data = parse_file(path)
        combined.queries.extend(data.queries)
        combined.pages.extend(data.pages)
        combined.query_pages.extend(data.query_pages)
        combined.source_files.extend(data.source_files)
    return combined


def validate(data: GSCData) -> list[str]:
    """Geeft een lijst van waarschuwingen als data onvolledig of leeg is."""
    warnings = []
    if not data.queries and not data.pages:
        warnings.append("Geen zoekterm- of paginadata gevonden. Controleer de kolomnamen in het bestand.")
    if not data.queries:
        warnings.append("Geen zoektermen gevonden — analyses op queryniveau niet mogelijk.")
    if not data.pages:
        warnings.append("Geen paginadata gevonden — analyses op paginaniveau niet mogelijk.")
    if not data.query_pages:
        warnings.append("Geen gecombineerde query+page data — cannibalisatiedetectie niet mogelijk.")
    return warnings
