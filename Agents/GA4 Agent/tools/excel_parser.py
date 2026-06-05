"""
Parser voor Google Analytics 4 Excel/CSV exports.

Ondersteunde exportformaten:
- GA4 UI Export > Pagina's (pagepath, sessions, pageviews, bounce rate, etc.)
- GA4 UI Export > Kanalen (channel group, sessions, users, etc.)
- Looker Studio exports met dezelfde kolomnamen
- Zowel .xlsx als .csv, EN en NL kolomnamen

Verwachte sheets (auto-detect op kolomnamen):
  - 'pages'    — paginaprestaties (pagepath + sessions + ...)
  - 'channels' — kanaalverdeling (channel + sessions + ...)
  - 'landing'  — landingspaginaprestaties (landingpage/pagepath + sessions + ...)
  - 'devices'  — apparaatuitspitsing (device + sessions + ...)
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Datamodellen
# ---------------------------------------------------------------------------

@dataclass
class PageRow:
    path: str
    title: str
    sessions: int
    pageviews: int
    bounce_rate: float            # percentage (bijv. 68.3)
    avg_engagement_seconds: float
    conversions: int


@dataclass
class ChannelRow:
    channel: str
    sessions: int
    users: int
    pageviews: int
    bounce_rate: float
    avg_session_duration_s: float
    conversions: int


@dataclass
class LandingPageRow:
    path: str
    sessions: int
    users: int
    bounce_rate: float
    avg_session_duration_s: float
    conversions: int


@dataclass
class DeviceRow:
    device: str
    sessions: int
    users: int
    bounce_rate: float
    conversions: int


@dataclass
class GA4Data:
    """Container voor alle geparsede GA4-data uit één of meerdere bestanden."""
    pages: list[PageRow] = field(default_factory=list)
    channels: list[ChannelRow] = field(default_factory=list)
    landing_pages: list[LandingPageRow] = field(default_factory=list)
    devices: list[DeviceRow] = field(default_factory=list)
    source_files: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Kolom-normalisatie
# ---------------------------------------------------------------------------

_COL_MAP: dict[str, str] = {
    # Pagina / landingspagina
    "page path": "path",
    "page path and screen class": "path",
    "pagepath": "path",
    "page": "path",
    "pagina": "path",
    "pagina-pad": "path",
    "landing page": "path",
    "landingpage": "path",
    "landingspagina": "path",
    "landing page + query string": "path",
    # Paginatitel
    "page title": "title",
    "page title and screen name": "title",
    "paginatitel": "title",
    "title": "title",
    # Kanaal
    "session default channel group": "channel",
    "default channel group": "channel",
    "channel": "channel",
    "kanaal": "channel",
    "session medium": "channel",
    # Apparaat
    "device category": "device",
    "device": "device",
    "apparaat": "device",
    # Sessies
    "sessions": "sessions",
    "sessies": "sessions",
    # Gebruikers
    "total users": "users",
    "users": "users",
    "active users": "users",
    "gebruikers": "users",
    # Pageviews
    "views": "pageviews",
    "pageviews": "pageviews",
    "screen page views": "pageviews",
    "paginaweergaven": "pageviews",
    # Bounce rate
    "bounce rate": "bounce_rate",
    "bouncepercentage": "bounce_rate",
    "bounceRate": "bounce_rate",
    # Gemiddelde sessieduur / engagementtijd
    "average session duration": "avg_duration",
    "avg. session duration": "avg_duration",
    "gemiddelde sessieduur": "avg_duration",
    "average engagement time": "avg_duration",
    "average engagement time per session": "avg_duration",
    "gem. betrokkenheidstijd per sessie": "avg_duration",
    "gem. sessieduur": "avg_duration",
    # Conversies
    "conversions": "conversions",
    "key events": "conversions",
    "key events total": "conversions",
    "conversies": "conversions",
    "doelconversies": "conversions",
}


def _normalize_header(raw: str) -> str:
    cleaned = str(raw).strip().lower()
    return _COL_MAP.get(cleaned, cleaned)


def _parse_float(value: str) -> float:
    v = str(value).strip().replace(",", ".").replace("%", "").replace("\xa0", "").strip()
    return float(v) if v and v not in ("-", "—", "n/a") else 0.0


def _parse_bounce(value: str) -> float:
    """Bounce rate: zet om naar percentage. GA4 geeft soms 0.683, soms 68.3%."""
    raw = str(value).strip().replace(",", ".").replace("%", "").replace("\xa0", "").strip()
    if not raw or raw in ("-", "—", "n/a"):
        return 0.0
    f = float(raw)
    return round(f * 100, 1) if f < 1.0 else round(f, 1)


def _parse_duration(value: str) -> float:
    """
    Sessieduur: accepteert seconden (bijv. '123.4') of mm:ss formaat ('2:03').
    Retourneert altijd seconden als float.
    """
    v = str(value).strip().replace(",", ".").replace("\xa0", "").strip()
    if not v or v in ("-", "—", "n/a"):
        return 0.0
    if ":" in v:
        parts = v.split(":")
        try:
            if len(parts) == 2:
                return float(parts[0]) * 60 + float(parts[1])
            if len(parts) == 3:
                return float(parts[0]) * 3600 + float(parts[1]) * 60 + float(parts[2])
        except ValueError:
            return 0.0
    try:
        return float(v)
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# Sheet-type detectie
# ---------------------------------------------------------------------------

def _detect_sheet_type(keys: set[str]) -> str:
    """Detecteer het type data op basis van aanwezige genormaliseerde kolomnamen."""
    has_path = "path" in keys
    has_channel = "channel" in keys
    has_device = "device" in keys

    if has_device and not has_channel and not has_path:
        return "devices"
    if has_channel and not has_path:
        return "channels"
    if has_path:
        return "pages"  # pages én landing pages beide via 'path'
    return "unknown"


# ---------------------------------------------------------------------------
# Rijen lezen
# ---------------------------------------------------------------------------

def _rows_from_csv(path: Path) -> list[dict]:
    with open(path, encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        return [
            {_normalize_header(k): v for k, v in row.items()}
            for row in reader
        ]


def _rows_from_sheet(sheet) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    # Sla lege koprijen over (GA4 exports hebben soms 1-2 lege rijen bovenaan)
    header_idx = 0
    for i, row in enumerate(rows):
        if any(cell is not None for cell in row):
            header_idx = i
            break
    headers = [
        _normalize_header(str(rows[header_idx][i]) if rows[header_idx][i] is not None else "")
        for i in range(len(rows[header_idx]))
    ]
    result = []
    for row in rows[header_idx + 1:]:
        if all(cell is None for cell in row):
            continue
        result.append({
            headers[i]: (str(row[i]) if row[i] is not None else "")
            for i in range(len(headers))
        })
    return result


# ---------------------------------------------------------------------------
# Omzetten naar dataclasses
# ---------------------------------------------------------------------------

def _to_page_rows(rows: list[dict]) -> list[PageRow]:
    result = []
    for r in rows:
        path = r.get("path", "").strip()
        if not path:
            continue
        try:
            result.append(PageRow(
                path=path,
                title=r.get("title", "").strip(),
                sessions=int(_parse_float(r.get("sessions", "0"))),
                pageviews=int(_parse_float(r.get("pageviews", "0"))),
                bounce_rate=_parse_bounce(r.get("bounce_rate", "0")),
                avg_engagement_seconds=_parse_duration(r.get("avg_duration", "0")),
                conversions=int(_parse_float(r.get("conversions", "0"))),
            ))
        except (ValueError, TypeError):
            continue
    return result


def _to_landing_page_rows(rows: list[dict]) -> list[LandingPageRow]:
    result = []
    for r in rows:
        path = r.get("path", "").strip()
        if not path:
            continue
        try:
            result.append(LandingPageRow(
                path=path,
                sessions=int(_parse_float(r.get("sessions", "0"))),
                users=int(_parse_float(r.get("users", "0"))),
                bounce_rate=_parse_bounce(r.get("bounce_rate", "0")),
                avg_session_duration_s=_parse_duration(r.get("avg_duration", "0")),
                conversions=int(_parse_float(r.get("conversions", "0"))),
            ))
        except (ValueError, TypeError):
            continue
    return result


def _to_channel_rows(rows: list[dict]) -> list[ChannelRow]:
    result = []
    for r in rows:
        channel = r.get("channel", "").strip()
        if not channel:
            continue
        try:
            result.append(ChannelRow(
                channel=channel,
                sessions=int(_parse_float(r.get("sessions", "0"))),
                users=int(_parse_float(r.get("users", "0"))),
                pageviews=int(_parse_float(r.get("pageviews", "0"))),
                bounce_rate=_parse_bounce(r.get("bounce_rate", "0")),
                avg_session_duration_s=_parse_duration(r.get("avg_duration", "0")),
                conversions=int(_parse_float(r.get("conversions", "0"))),
            ))
        except (ValueError, TypeError):
            continue
    return result


def _to_device_rows(rows: list[dict]) -> list[DeviceRow]:
    result = []
    for r in rows:
        device = r.get("device", "").strip()
        if not device:
            continue
        try:
            result.append(DeviceRow(
                device=device,
                sessions=int(_parse_float(r.get("sessions", "0"))),
                users=int(_parse_float(r.get("users", "0"))),
                bounce_rate=_parse_bounce(r.get("bounce_rate", "0")),
                conversions=int(_parse_float(r.get("conversions", "0"))),
            ))
        except (ValueError, TypeError):
            continue
    return result


# ---------------------------------------------------------------------------
# Publieke API
# ---------------------------------------------------------------------------

def parse_file(path: str | Path) -> GA4Data:
    """
    Parseer één Excel of CSV bestand met GA4 exportdata.

    Ondersteunt:
    - .csv  — één sheet (auto-detect type)
    - .xlsx — één of meerdere sheets (auto-detect per sheet)
    """
    p = Path(path)
    data = GA4Data(source_files=[str(p)])

    if p.suffix.lower() == ".csv":
        rows = _rows_from_csv(p)
        if not rows:
            return data
        keys = set(rows[0].keys())
        sheet_type = _detect_sheet_type(keys)
        _apply_rows(data, rows, sheet_type)

    elif p.suffix.lower() in (".xlsx", ".xls"):
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl is vereist voor .xlsx bestanden: pip install openpyxl")
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = _rows_from_sheet(ws)
            if not rows:
                continue
            keys = set(rows[0].keys())
            sheet_type = _detect_sheet_type(keys)
            # Hinteer op sheetname als auto-detect onduidelijk is
            if sheet_type == "pages":
                name_lower = sheet_name.lower()
                if any(w in name_lower for w in ("landing", "entry", "ingang")):
                    sheet_type = "landing"
            _apply_rows(data, rows, sheet_type)
        wb.close()
    else:
        raise ValueError(f"Niet-ondersteund bestandstype: {p.suffix}. Gebruik .csv of .xlsx")

    return data


def _apply_rows(data: GA4Data, rows: list[dict], sheet_type: str) -> None:
    if sheet_type == "pages":
        data.pages.extend(_to_page_rows(rows))
    elif sheet_type == "landing":
        data.landing_pages.extend(_to_landing_page_rows(rows))
    elif sheet_type == "channels":
        data.channels.extend(_to_channel_rows(rows))
    elif sheet_type == "devices":
        data.devices.extend(_to_device_rows(rows))


def parse_files(paths: list[str | Path]) -> GA4Data:
    """
    Parseer meerdere GA4-exportbestanden en combineer ze tot één GA4Data object.
    Handig als pagina's, kanalen en devices als aparte exports zijn opgeslagen.
    """
    combined = GA4Data()
    for path in paths:
        d = parse_file(path)
        combined.pages.extend(d.pages)
        combined.channels.extend(d.channels)
        combined.landing_pages.extend(d.landing_pages)
        combined.devices.extend(d.devices)
        combined.source_files.extend(d.source_files)
    return combined


def validate(data: GA4Data) -> list[str]:
    """Geeft een lijst van waarschuwingen als data onvolledig of leeg is."""
    warnings = []
    if not data.pages and not data.channels:
        warnings.append("Geen pagina- of kanaaldata gevonden. Controleer de kolomnamen in het bestand.")
    if not data.pages:
        warnings.append("Geen paginadata — analyses op paginaniveau niet mogelijk.")
    if not data.channels:
        warnings.append("Geen kanaaldata — kanaalanalyse niet mogelijk.")
    if not data.landing_pages:
        warnings.append("Geen landingspaginadata — landingspagina-analyse niet mogelijk.")
    if not data.devices:
        warnings.append("Geen apparaatdata — device-analyse niet mogelijk.")
    return warnings
